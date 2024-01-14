from django.shortcuts import render
from openpyxl import load_workbook
from .models import MarketServicePrice

def upload_contract(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            # TODO - Clean data before saving to database
            outlet, site, stream_id, service, type, qty_scheduled, frequency, unit_price, start_date, end_date, stage, route_schedule, days = row
            MarketServicePrice.objects.create(
                outlet=outlet,
                site=site,
                stream_id=stream_id,
                service=service,
                type=type,
                qty_scheduled=qty_scheduled,
                frequency=frequency,
                unit_price=unit_price,
                start_date=start_date,
                end_date=end_date,
                stage=stage,
                route_schedule=route_schedule,
                days=days
            )
        return render(request, 'upload_success.html')

    return render(request, 'contracts.html')
